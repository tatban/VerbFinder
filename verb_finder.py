from __future__ import unicode_literals, print_function
from argparse import ArgumentParser
from functools import partial
import re
import pdfplumber
import spacy
from spacy.language import Language
from spacy.tokens.doc import Doc
from spacy.tokens.span import Span
from openpyxl.workbook import Workbook
from openpyxl.styles import Font
from openpyxl.utils import exceptions as opex
from table_handler import get_bounding_boxes, table_filter
import time
from typing import Dict, List, Tuple
from pathlib import Path
import sys
import warnings
from utils import read_yaml


def text_cleaner(txt: str, patterns: List[str]) -> str:
    """
    Removes matching regex patterns from the input text.
    :param txt: input text
    :param patterns: list of regex patterns
    :return: clean text after removing the matching patterns
    """
    return re.sub('|'.join(patterns), "", txt)


def setup_nlp_pipeline(spacy_model: str, use_sentencizer: bool = True) -> Language:
    """
    Loads a pretrained spacy language model for performing the nlp pipeline tasks.
    :param spacy_model: name of the language model (eg. 'en_core_web_sm')
    :param use_sentencizer: whether to use sentence segmentation as a part of nlp pipeline
    :return: language specific pretrained spacy language model object
    """
    if not spacy.util.is_package(spacy_model):
        try:
            spacy.cli.download(spacy_model)
        except Exception as e:
            sys.exit(
                f"Couldn't download {spacy_model} model due to following "
                f"exception:\n{e}")
    nlp_pipeline = spacy.load(spacy_model)
    if use_sentencizer:
        nlp_pipeline.add_pipe('sentencizer')
    return nlp_pipeline


def get_verbs(sentence: Span, lemmatize: bool = True, detect_aux: bool = True) -> List[str]:
    """
    Returns the verbs of a input sentence. If lemmatize=True, verbs are converted to root
    forms. If detect_aux=True, auxiliary verbs (eg. 'be', 'can') are also detected.
    :param sentence:
    :param lemmatize: whether to convert the verbs to their root forms
    :param detect_aux: whether to detect auxiliary verbs like 'be', 'can'
    :return: list of verbs strings
    """
    verb_types = ["VERB", "AUX"] if detect_aux else ["VERB"]
    if lemmatize:
        return [word.lemma_ for word in sentence if word.pos_ in verb_types]
    else:
        return [word.text for word in sentence if word.pos_ in verb_types]


def process_document(document: Doc,
                     filter_single_word: bool = True,
                     lemmatize: bool = True,
                     detect_aux: bool = True) -> List[Tuple[str, List[str]]]:
    """
    Loops over all the sentences of the document and extracts verbs from those. Lists the
    sentences and corresponding verbs in a list of tuples
    :param document:
    :param filter_single_word: If true, filter out single word sentences (which are most
    likely to be the outcome of sentence segmentation errors.)
    :param lemmatize: whether to convert the verbs to their root forms
    :param detect_aux: whether to detect auxiliary verbs like 'be', 'can'
    :return:
    """
    if filter_single_word:
        return [(sentence.text.strip().replace('\n', ''),
                 get_verbs(sentence, lemmatize=lemmatize, detect_aux=detect_aux))
                for sentence in list(document.sents)
                if len(sentence.text.strip().replace('\n', '').split()) > 1]
    else:
        return [(sentence.text.strip().replace('\n', ''),
                 get_verbs(sentence, lemmatize=lemmatize, detect_aux=detect_aux))
                for sentence in list(document.sents)]


def save_xlsx(table: List[Tuple[str, List[str]]], out_path: str):
    """
    Saves the list of tuples of sentences and corresponding verbs in xlsx file
    :param table: list of tuples in the form [(sentence, [verbs])]
    :param out_path: string path of a writable .xlsx file. If the file exists already,
    it will be overwritten.
    """
    wb = Workbook()
    ws = wb.active  # grab the active worksheet
    ws['A1'] = "Sentences"
    ws['A1'].font = Font(bold=True)
    ws['B1'] = "Verb-entities"
    ws['B1'].font = Font(bold=True)
    for i, (sentence, verbs) in enumerate(table, 2):
        ws[f"A{i}"] = sentence
        ws[f"B{i}"] = ", ".join(verbs)
    try:
        wb.save(out_path)
        print(f"Extracted sentences and verbs are saved successfully in {out_path} file.\n")
    except opex as ope:
        sys.exit(f"Couldn't save file due to following exception:\n{ope.message}")


def driver(configuration: Dict,
           pdf_path: str,
           pages: List[int],
           out_path: str):
    """
    Driver function to extract the sentences and corresponding verbs from a pdf document
    :param configuration: a dictionary having settings for the program read from config.yaml
    :param pdf_path: file path (string) of the input pdf file
    :param pages: page range [start, end] containing the sections of interest. default is None,
    meaning all pages.
    :param out_path: string path of a writable .xlsx file. If the file exists already,
    it will be overwritten.
    """
    # validate input path
    assert pdf_path.split(".")[-1].lower() == "pdf", "Only pdf file is supported at the moment"
    assert Path(pdf_path).is_file(), f"{pdf_path} file doesn't exist. Please enter a valid " \
                                     f"pdf file path"

    # validate output path
    assert out_path.split(".")[-1].lower() == "xlsx", "output needs to be a .xlsx file"
    # warn user before overwriting the file if the file already exists
    if Path(out_path).is_file():
        warnings.warn(f"File {out_path} already exists. If you continue the file will be "
                      f"overwritten.\n")
        choice = input("Continue? [Y/n]")
        if choice.lower() in {'no', 'n'}:
            sys.exit("Aborting the execution not to overwrite the output file.\n")

    # validate page ranges
    if pages is not None:
        assert len(pages) == 2, "page range should be like [stat page umber, end page number]"
        assert pages[0] <= pages[1], "start page number must be less than end page number"
        assert pages[0] > 0, "lowest possible page number is 1"

    # open pdf
    try:
        pdf = pdfplumber.open(pdf_path)
    except Exception as e:
        sys.exit(f"Couldn't open pdf file {pdf_path} due to following exception:\n{e}")

    # handle if all pages need to be processed
    if pages is None:
        pages = [1, len(pdf.pages)]

    # keep nlp pipeline ready
    nlp = setup_nlp_pipeline(
        configuration.get('NLP_PIPELINE', {}).get('SPACY_MODEL', 'en_core_web_sm'),
        use_sentencizer=configuration.get('NLP_PIPELINE', {}).get('USE_SENTENCIZER', True)
    )

    # loop pages
    clean_text = ""
    for p in pdf.pages[pages[0] - 1:pages[1] - 1]:
        # find table locations if any
        bboxes = get_bounding_boxes(p)
        page_specific_table_filter = partial(table_filter, bboxes)
        # find the text not lying in the table region
        text = p.filter(page_specific_table_filter).extract_text()
        # clean text and append
        clean_text = '\n\n'.join([
            clean_text,
            text_cleaner(text,
                         configuration.get('TXT_CLEANER', {}).get('FILTER_PATTERNS', None))
        ])

    # run nlp pipeline
    doc = nlp(clean_text)
    table = process_document(
        doc,
        filter_single_word=configuration.get('NLP_PIPELINE', {}).get('FILTER_SINGLE_WORDS',
                                                                     True),
        lemmatize=configuration.get('NLP_PIPELINE', {}).get('LEMMATIZE', True),
        detect_aux=configuration.get('NLP_PIPELINE', {}).get('DETECT_AUX', True)
    )
    save_xlsx(table, out_path)


if __name__ == "__main__":
    """
    all default values for the arguments are chosen here for the specific task 
    of extracting sentences and corresponding verbs from section 6.3 in the OPC 
    Unified Architecture Documentation obtained after pdf conversion of the htm.
    Please look at readme.md to know more.)
    """
    parser = ArgumentParser()  # CLI argument parser object
    # adding arguments
    parser.add_argument("--cfg", type=str,
                        help="String path of the yaml config file (.yaml)",
                        default="config.yaml")
    parser.add_argument("--pdf", type=str,
                        help="String path of the input pdf file (.pdf)",
                        default="test_pdf.pdf")
    parser.add_argument("--pgs", type=List[int],
                        help="Pages range to process [start, end] counting from 1",
                        default=[18, 24])
    parser.add_argument("--out", type=str,
                        help="string path of output file (.xlsx)",
                        default="result.xlsx")
    args = parser.parse_args()  # parsing arguments
    config = read_yaml(args.cfg)
    start = time.time()
    driver(config, args.pdf, args.pgs, args.out)
    print(f"Time elapsed: {time.time() - start} seconds")
